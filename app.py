import streamlit as st
import pdfplumber
import hashlib
import json
import io
import csv
import os
from datetime import date
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from supabase import create_client, Client
from openai import OpenAI
from prompts import (
    SYS_PROMPT_ESTRUCTURACION, SYS_PROMPT_EVALUACION,
    get_user_prompt_estructuracion, get_user_prompt_evaluacion
)

# ==========================================
# 1. CONFIGURACIÓN E INICIALIZACIÓN
# ==========================================
st.set_page_config(page_title="AI Resume Screener", page_icon="🚀", layout="wide")

@st.cache_resource
def init_connections():
    """Inicializa Supabase y el cliente LLM (GitHub Models) a prueba de fallos."""
    try:
        # Intento 1: Streamlit Cloud Secrets
        supa_url = st.secrets["SUPABASE_URL"]
        supa_key = st.secrets["SUPABASE_KEY"]
        github_token = st.secrets["GITHUB_TOKEN"]
    except (FileNotFoundError, KeyError):
        # Intento 2: Entorno local (.env)
        load_dotenv()

        supa_url = os.environ.get("SUPABASE_URL")
        supa_key = os.environ.get("SUPABASE_KEY")
        github_token = os.environ.get("GITHUB_TOKEN")

    if not supa_url or not supa_key or not github_token:
        st.error("🚨 Faltan credenciales. Verifica Supabase URL/KEY y GITHUB_TOKEN.")
        st.stop()

    db_client = create_client(supa_url, supa_key)
    llm_client = OpenAI(
        base_url="https://models.inference.ai.azure.com",
        api_key=github_token,
    )

    return db_client, llm_client

supabase, llm = init_connections()
MODELO_NUBE = "gpt-4.1"

# Precios gpt-4.1 (USD por 1M tokens)
_PRECIO_INPUT        = 2.50
_PRECIO_INPUT_CACHE  = 1.25
_PRECIO_OUTPUT       = 10.00

def calcular_costo(tokens_input: int, tokens_output: int, tokens_cached: int = 0) -> float:
    """Calcula el costo en USD considerando tokens en caché."""
    tokens_no_cache = tokens_input - tokens_cached
    return (
        tokens_no_cache * _PRECIO_INPUT / 1_000_000
        + tokens_cached  * _PRECIO_INPUT_CACHE / 1_000_000
        + tokens_output  * _PRECIO_OUTPUT / 1_000_000
    )

# ==========================================
# 2. FUNCIONES CORE (NLP & LLM)
# ==========================================
def extraer_texto_en_memoria(archivo_pdf) -> tuple[str, int]:
    """Extrae texto y número de páginas del PDF desde memoria RAM."""
    texto = ""
    num_paginas = 0
    try:
        with pdfplumber.open(archivo_pdf) as pdf:
            num_paginas = len(pdf.pages)
            for pagina in pdf.pages:
                extraido = pagina.extract_text()
                if extraido:
                    texto += extraido + "\n"
        return texto.strip(), num_paginas
    except Exception as e:
        st.error(f"Error al leer el PDF: {e}")
        return "", 0

def interactuar_con_gpt(prompt: str, rol_sistema: str) -> tuple[dict, object]:
    """Habla con GPT-4.1 vía GitHub y devuelve (json_dict, usage)."""
    try:
        response = llm.chat.completions.create(
            messages=[
                {"role": "system", "content": rol_sistema},
                {"role": "user", "content": prompt}
            ],
            model=MODELO_NUBE,
            temperature=0,
            response_format={"type": "json_object"}
        )
        contenido = response.choices[0].message.content
        return json.loads(contenido), response.usage
    except Exception as e:
        st.error(f"Error en la inferencia con GPT-4.1: {e}")
        return {}, None

# ==========================================
# 3. FUNCIONES DE VISUALIZACIÓN
# ==========================================

DIM_LABELS = ["Skills técnicos", "Experiencia", "Educación", "Idiomas", "Fit general"]
DIM_KEYS   = ["score_skills_tecnicos", "score_experiencia", "score_educacion", "score_idiomas", "score_fit_general"]

DECISION_COLOR_HEX = {
    "prioridad":   "#22c55e",
    "entrevistar": "#eab308",
    "considerar":  "#f97316",
    "descartar":   "#ef4444",
}

def radar_candidato(evaluacion: dict, nombre: str) -> go.Figure:
    valores = [evaluacion.get(k) or 0 for k in DIM_KEYS]
    valores_cerrados = valores + [valores[0]]
    labels_cerrados  = DIM_LABELS + [DIM_LABELS[0]]
    rec   = evaluacion.get("recomendacion", "descartar")
    color = DECISION_COLOR_HEX.get(rec, "#6b7280")

    fig = go.Figure(go.Scatterpolar(
        r=valores_cerrados,
        theta=labels_cerrados,
        fill="toself",
        fillcolor=color,
        line=dict(color=color),
        opacity=0.6,
        name=nombre,
    ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        showlegend=False,
        margin=dict(t=30, b=30, l=40, r=40),
        height=320,
    )
    return fig


def donut_decisiones(filas: list) -> go.Figure:
    conteo = {"Prioridad": 0, "Entrevistar": 0, "Considerar": 0, "Descartar": 0}
    for f in filas:
        decision = f.get("recomendacion", "descartar").capitalize()
        if decision in conteo:
            conteo[decision] += 1
    colores = [
        DECISION_COLOR_HEX["prioridad"],
        DECISION_COLOR_HEX["entrevistar"],
        DECISION_COLOR_HEX["considerar"],
        DECISION_COLOR_HEX["descartar"],
    ]
    fig = go.Figure(go.Pie(
        labels=list(conteo.keys()),
        values=list(conteo.values()),
        hole=0.55,
        marker_colors=colores,
        textinfo="label+percent",
    ))
    fig.update_layout(
        showlegend=False,
        margin=dict(t=20, b=20, l=20, r=20),
        height=300,
    )
    return fig


def histograma_scores(filas: list) -> go.Figure:
    scores = [f["score_total"] for f in filas if f.get("score_total") is not None]
    fig = go.Figure(go.Histogram(
        x=scores,
        xbins=dict(start=0, end=100, size=10),
        marker_color="#6366f1",
        opacity=0.8,
    ))
    fig.update_layout(
        xaxis=dict(title="Score total", range=[0, 100]),
        yaxis=dict(title="Candidatos"),
        bargap=0.1,
        margin=dict(t=20, b=40, l=40, r=20),
        height=300,
    )
    return fig


def radar_promedio(filas: list) -> go.Figure:
    promedios = []
    for key in DIM_KEYS:
        vals = [f[key] for f in filas if f.get(key) is not None]
        promedios.append(sum(vals) / len(vals) if vals else 0)

    valores_cerrados = promedios + [promedios[0]]
    labels_cerrados  = DIM_LABELS + [DIM_LABELS[0]]

    fig = go.Figure(go.Scatterpolar(
        r=valores_cerrados,
        theta=labels_cerrados,
        fill="toself",
        fillcolor="#6366f1",
        line=dict(color="#6366f1"),
        opacity=0.5,
    ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        showlegend=False,
        margin=dict(t=30, b=30, l=40, r=40),
        height=340,
    )
    return fig


def bar_ranking(resultados: list) -> go.Figure:
    nombres   = [r["cv_json"].get("nombre_candidato", "—") for r in resultados]
    scores    = [r["evaluacion"].get("score_total", 0) for r in resultados]
    decisiones = [r["evaluacion"].get("recomendacion", "descartar") for r in resultados]
    colores   = [DECISION_COLOR_HEX.get(d, "#6b7280") for d in decisiones]

    fig = go.Figure(go.Bar(
        x=scores,
        y=nombres,
        orientation="h",
        marker_color=colores,
        text=[f"{s}/100" for s in scores],
        textposition="outside",
    ))
    fig.update_layout(
        xaxis=dict(range=[0, 110], title="Score"),
        yaxis=dict(autorange="reversed"),
        margin=dict(t=20, b=20, l=10, r=60),
        height=max(200, len(resultados) * 55),
    )
    return fig


# ==========================================
# 4. FUNCIONES DE EXPORTACIÓN
# ==========================================

DECISION_ES = {
    "prioridad":   "Priorizar",
    "entrevistar": "Entrevistar",
    "considerar":  "Considerar",
    "descartar":   "Descartar",
}

MESES_ES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo",    "04": "Abril",
    "05": "Mayo",  "06": "Junio",   "07": "Julio",    "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}

def mes_legible(fecha_str: str) -> str:
    """Convierte '2026-04' o '2026-04-06' a 'Abril 2026'."""
    if len(fecha_str) >= 7:
        anio, mes = fecha_str[:4], fecha_str[5:7]
        return f"{MESES_ES.get(mes, mes)} {anio}"
    return fecha_str


def generar_excel(filas_export: list) -> bytes:
    """
    Genera un archivo Excel limpio para el reclutador.
    filas_export: lista de dicts con campos estandarizados.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos evaluados"

    FILL_DECISION = {
        "Priorizar":   PatternFill("solid", fgColor="22c55e"),
        "Entrevistar": PatternFill("solid", fgColor="eab308"),
        "Considerar":  PatternFill("solid", fgColor="f97316"),
        "Descartar":   PatternFill("solid", fgColor="ef4444"),
    }

    cabeceras = [
        "Proceso de selección", "Candidato", "Último cargo", "Empresa",
        "Puntuación total", "Skills técnicos", "Experiencia", "Educación",
        "Idiomas", "Encaje general", "Decisión recomendada",
        "Por qué la IA tomó esta decisión", "Fecha de evaluación",
    ]
    for col, texto in enumerate(cabeceras, 1):
        celda = ws.cell(row=1, column=col, value=texto)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(wrap_text=True)

    for fila_i, f in enumerate(filas_export, 2):
        decision_label = DECISION_ES.get(f.get("recomendacion", "descartar"), "Descartar")
        valores = [
            f.get("proceso_nombre", "—"),
            f.get("nombre_candidato", "—"),
            f.get("ultimo_cargo", "—"),
            f.get("ultima_empresa", "—"),
            f.get("score_total"),
            f.get("score_skills_tecnicos"),
            f.get("score_experiencia"),
            f.get("score_educacion"),
            f.get("score_idiomas"),
            f.get("score_fit_general"),
            decision_label,
            f.get("justificacion_general", ""),
            f.get("fecha", ""),
        ]
        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=fila_i, column=col, value=val)
            if col == 11:  # columna "Decisión"
                fill = FILL_DECISION.get(decision_label)
                if fill:
                    celda.fill = fill
            if col == 12:  # justificación — ajuste de texto
                celda.alignment = Alignment(wrap_text=True)

    # Ancho de columnas razonable
    anchos = [28, 24, 24, 22, 14, 14, 14, 14, 10, 14, 20, 60, 16]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_html_imprimible(resultados: list, proceso_nombre: str) -> str:
    """Genera un HTML autocontenido listo para imprimir o guardar como PDF desde el navegador."""
    filas_html = ""
    for pos, r in enumerate(resultados, 1):
        cv = r["cv_json"]
        ev = r["evaluacion"]
        rec = ev.get("recomendacion", "descartar")
        decision_label = DECISION_ES.get(rec, "—")
        color = DECISION_COLOR_HEX.get(rec, "#6b7280")
        fortalezas = ev.get("fortalezas", [])
        puntos = "".join(f"<li>{f}</li>" for f in fortalezas)
        filas_html += f"""
        <tr>
          <td style="text-align:center;font-weight:bold;">{pos}</td>
          <td>{cv.get("nombre_candidato","—")}</td>
          <td>{cv.get("ultimo_cargo","—") or "—"}</td>
          <td style="text-align:center;font-weight:bold;">{ev.get("score_total",0):.0f}/100</td>
          <td style="text-align:center;">
            <span style="background:{color};color:white;padding:3px 10px;
                         border-radius:12px;font-size:12px;">{decision_label}</span>
          </td>
          <td><ul style="margin:0;padding-left:16px;font-size:12px;">{puntos}</ul></td>
          <td style="font-size:12px;color:#444;">{ev.get("justificacion_general","")}</td>
        </tr>"""

    hoy = date.today().strftime("%d/%m/%Y")
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <title>Ranking — {proceso_nombre}</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 32px; color: #222; }}
    h1   {{ font-size: 20px; margin-bottom: 4px; }}
    p.meta {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th  {{ background: #1e293b; color: white; padding: 8px 12px; text-align: left; }}
    td  {{ padding: 8px 12px; border-bottom: 1px solid #e2e8f0; vertical-align: top; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    @media print {{ body {{ margin: 16px; }} button {{ display:none; }} }}
  </style>
</head>
<body>
  <h1>Ranking de candidatos — {proceso_nombre}</h1>
  <p class="meta">Generado el {hoy} &nbsp;·&nbsp; {len(resultados)} candidatos evaluados</p>
  <table>
    <thead>
      <tr>
        <th>#</th><th>Candidato</th><th>Último cargo</th>
        <th>Puntuación</th><th>Decisión</th>
        <th>Puntos fuertes</th><th>Por qué esta decisión</th>
      </tr>
    </thead>
    <tbody>{filas_html}</tbody>
  </table>
</body>
</html>"""


# ==========================================
# 5. INTERFAZ DE USUARIO
# ==========================================
st.title("☁️ Sistema de Selección Automatizada")
st.markdown("Sube un CV, define el Job Spec y deja que la IA evalúe la compatibilidad con explicabilidad total.")

# --- Session state defaults ---
if "proceso_nombre" not in st.session_state:
    st.session_state.proceso_nombre = ""
if "job_spec" not in st.session_state:
    st.session_state.job_spec = ""
if "resultados_sesion" not in st.session_state:
    st.session_state.resultados_sesion = []  # lista de {cv_json, evaluacion} por candidato

tab_config, tab_evaluar, tab_historial = st.tabs([
    "⚙️ Configuración del proceso",
    "📄 Evaluar CV",
    "📊 Historial",
])

# ==========================================
# TAB 1 — CONFIGURACIÓN DEL PROCESO
# ==========================================
with tab_config:
    st.subheader("Configuración del proceso de selección")
    st.markdown("Define el contexto del proceso. Estos datos se usarán en la pestaña **Evaluar CV** y quedan activos durante toda la sesión.")

    proceso_nombre_input = st.text_input(
        "Nombre del proceso de selección",
        value=st.session_state.proceso_nombre,
        placeholder="Ej: Data Engineer Q1 2026",
    )

    job_spec_input = st.text_area(
        "Requerimientos del Puesto (Job Spec)",
        value=st.session_state.job_spec,
        height=250,
        placeholder="Ej: Buscamos un Data Scientist con 3 años de experiencia en Python...",
    )

    if st.button("Guardar configuración", type="primary"):
        if not proceso_nombre_input or not job_spec_input:
            st.warning("⚠️ Completa el nombre del proceso y el Job Spec antes de guardar.")
        else:
            st.session_state.proceso_nombre = proceso_nombre_input
            st.session_state.job_spec = job_spec_input
            st.success("✅ Configuración guardada. Ya puedes evaluar CVs.")

# ==========================================
# TAB 2 — EVALUAR CV
# ==========================================
with tab_evaluar:
    st.subheader("Evaluación de candidato")

    if not st.session_state.proceso_nombre or not st.session_state.job_spec:
        st.info("ℹ️ Primero configura el proceso en la pestaña **⚙️ Configuración del proceso**.")
    else:
        st.markdown(
            f"**Proceso activo:** {st.session_state.proceso_nombre}  \n"
            f"**Job Spec:** {st.session_state.job_spec[:120]}{'...' if len(st.session_state.job_spec) > 120 else ''}"
        )
        st.divider()

        archivos_subidos = st.file_uploader(
            "Sube los CVs de los candidatos (PDF)",
            type=["pdf"],
            accept_multiple_files=True,
        )

        col_btn, col_clear = st.columns([3, 1])
        ejecutar = col_btn.button("Ejecutar Motor de Evaluación", type="primary", use_container_width=True)
        if col_clear.button("Limpiar resultados", use_container_width=True):
            st.session_state.resultados_sesion = []
            st.rerun()

        if ejecutar:
            if not archivos_subidos:
                st.warning("⚠️ Sube al menos un archivo PDF para continuar.")
            else:
                nuevos_resultados = []
                total = len(archivos_subidos)

                for i, archivo in enumerate(archivos_subidos, start=1):
                    with st.status(f"Procesando {i}/{total}: {archivo.name}", expanded=True) as status:
                        st.write("📄 Extrayendo texto del PDF...")
                        texto_cv, num_paginas = extraer_texto_en_memoria(archivo)

                        if not texto_cv:
                            status.update(label=f"❌ Error extrayendo {archivo.name}", state="error")
                            continue

                        st.write("🧠 Estructurando CV con GPT-4.1...")
                        prompt_estructura = get_user_prompt_estructuracion(texto_cv)
                        cv_json, usage_struct = interactuar_con_gpt(prompt_estructura, SYS_PROMPT_ESTRUCTURACION)

                        st.write("⚖️ Evaluando compatibilidad con el Job Spec...")
                        prompt_evaluacion = get_user_prompt_evaluacion(json.dumps(cv_json), st.session_state.job_spec)
                        evaluacion, usage_eval = interactuar_con_gpt(prompt_evaluacion, SYS_PROMPT_EVALUACION)

                        # Agregación de tokens de ambas llamadas
                        def _tokens(usage, attr):
                            return getattr(usage, attr, 0) or 0

                        def _cached(usage):
                            details = getattr(usage, "prompt_tokens_details", None)
                            return getattr(details, "cached_tokens", 0) or 0

                        tokens_input  = _tokens(usage_struct, "prompt_tokens")     + _tokens(usage_eval, "prompt_tokens")
                        tokens_output = _tokens(usage_struct, "completion_tokens") + _tokens(usage_eval, "completion_tokens")
                        tokens_cached = _cached(usage_struct) + _cached(usage_eval)
                        costo_usd     = calcular_costo(tokens_input, tokens_output, tokens_cached)

                        st.write("💾 Guardando en Supabase...")
                        archivo.seek(0)
                        pdf_bytes = archivo.read()
                        file_hash = hashlib.sha256(pdf_bytes).hexdigest()

                        resp_bronze = supabase.schema("bronze").table("raw_cvs").insert({
                            "proceso_nombre": st.session_state.proceso_nombre,
                            "filename":       archivo.name,
                            "file_hash":      file_hash,
                            "texto_crudo":    texto_cv,
                            "num_paginas":    num_paginas,
                            "tamanio_bytes":  len(pdf_bytes),
                        }).execute()
                        raw_cv_id = resp_bronze.data[0]["id"]

                        resp_silver = supabase.schema("silver").table("cv_estructurados").insert({
                            "raw_cv_id":             raw_cv_id,
                            "nombre_candidato":      cv_json.get("nombre_candidato"),
                            "email":                 cv_json.get("email"),
                            "telefono":              cv_json.get("telefono"),
                            "ubicacion":             cv_json.get("ubicacion"),
                            "resumen_perfil":        cv_json.get("resumen_perfil"),
                            "experiencia_anios":     cv_json.get("experiencia_anios"),
                            "ultimo_cargo":          cv_json.get("ultimo_cargo"),
                            "ultima_empresa":        cv_json.get("ultima_empresa"),
                            "educacion_nivel":       cv_json.get("educacion_nivel"),
                            "educacion_carrera":     cv_json.get("educacion_carrera"),
                            "educacion_institucion": cv_json.get("educacion_institucion"),
                            "skills_tecnicos":       cv_json.get("skills_tecnicos", []),
                            "idiomas":               cv_json.get("idiomas", []),
                            "experiencia_detalle":   cv_json.get("experiencia_detalle", []),
                            "educacion_detalle":     cv_json.get("educacion_detalle", []),
                        }).execute()
                        cv_estructurado_id = resp_silver.data[0]["id"]

                        supabase.schema("gold").table("evaluaciones").insert({
                            "cv_estructurado_id":    cv_estructurado_id,
                            "score_total":           evaluacion.get("score_total", 0),
                            "score_skills_tecnicos": evaluacion.get("score_skills_tecnicos"),
                            "score_experiencia":     evaluacion.get("score_experiencia"),
                            "score_educacion":       evaluacion.get("score_educacion"),
                            "score_idiomas":         evaluacion.get("score_idiomas"),
                            "score_fit_general":     evaluacion.get("score_fit_general"),
                            "recomendacion":         evaluacion.get("recomendacion", "descartar"),
                            "justificacion_general": evaluacion.get("justificacion_general"),
                            "fortalezas":            evaluacion.get("fortalezas", []),
                            "brechas":               evaluacion.get("brechas", []),
                            "tokens_input":          tokens_input,
                            "tokens_output":         tokens_output,
                            "costo_usd":             round(costo_usd, 6),
                        }).execute()

                        nuevos_resultados.append({"cv_json": cv_json, "evaluacion": evaluacion})
                        status.update(label=f"✅ {archivo.name}", state="complete", expanded=False)

                st.session_state.resultados_sesion = nuevos_resultados
                st.success(f"✅ {len(nuevos_resultados)}/{total} CVs procesados y persistidos.")

        # --- Ranking + detalle por candidato ---
        if st.session_state.resultados_sesion:
            resultados = sorted(
                st.session_state.resultados_sesion,
                key=lambda r: r["evaluacion"].get("score_total", 0),
                reverse=True,
            )

            st.divider()
            st.markdown("## Ranking de candidatos")

            DECISION_COLOR = {
                "prioridad":   "🟢",
                "entrevistar": "🟡",
                "considerar":  "🟠",
                "descartar":   "🔴",
            }

            ranking_filas = []
            export_filas  = []
            for pos, r in enumerate(resultados, start=1):
                ev  = r["evaluacion"]
                cv  = r["cv_json"]
                rec = ev.get("recomendacion", "descartar")
                ranking_filas.append({
                    "#":           pos,
                    "Candidato":   cv.get("nombre_candidato", "—"),
                    "Puntuación":  ev.get("score_total", 0),
                    "Skills":      ev.get("score_skills_tecnicos", "—"),
                    "Experiencia": ev.get("score_experiencia", "—"),
                    "Educación":   ev.get("score_educacion", "—"),
                    "Idiomas":     ev.get("score_idiomas", "—"),
                    "Encaje":      ev.get("score_fit_general", "—"),
                    "Decisión":    f"{DECISION_COLOR.get(rec, '')} {DECISION_ES.get(rec, rec.capitalize())}",
                })
                export_filas.append({
                    "proceso_nombre":        st.session_state.proceso_nombre,
                    "nombre_candidato":      cv.get("nombre_candidato", "—"),
                    "ultimo_cargo":          cv.get("ultimo_cargo", "—"),
                    "ultima_empresa":        cv.get("ultima_empresa", "—"),
                    "score_total":           ev.get("score_total"),
                    "score_skills_tecnicos": ev.get("score_skills_tecnicos"),
                    "score_experiencia":     ev.get("score_experiencia"),
                    "score_educacion":       ev.get("score_educacion"),
                    "score_idiomas":         ev.get("score_idiomas"),
                    "score_fit_general":     ev.get("score_fit_general"),
                    "recomendacion":         rec,
                    "justificacion_general": ev.get("justificacion_general", ""),
                    "fecha":                 date.today().strftime("%Y-%m-%d"),
                })

            st.dataframe(ranking_filas, use_container_width=True, hide_index=True)
            st.plotly_chart(bar_ranking(resultados), use_container_width=True)

            # --- Exportación ---
            nombre_proceso = st.session_state.proceso_nombre.replace(" ", "_")
            exp_col1, exp_col2, exp_col3 = st.columns(3)

            # CSV
            buf_csv = io.StringIO()
            writer = csv.DictWriter(buf_csv, fieldnames=export_filas[0].keys())
            writer.writeheader()
            writer.writerows(export_filas)
            exp_col1.download_button(
                label="⬇️ Descargar CSV",
                data=buf_csv.getvalue().encode("utf-8"),
                file_name=f"ranking_{nombre_proceso}.csv",
                mime="text/csv",
                help="Archivo de texto separado por comas, compatible con Excel y Google Sheets.",
            )

            # Excel
            exp_col2.download_button(
                label="📊 Descargar Excel",
                data=generar_excel(export_filas),
                file_name=f"ranking_{nombre_proceso}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Hoja de cálculo con colores por decisión y justificaciones.",
            )

            # HTML imprimible
            html_ranking = generar_html_imprimible(resultados, st.session_state.proceso_nombre)
            exp_col3.download_button(
                label="🖨️ Descargar para reunión (PDF)",
                data=html_ranking.encode("utf-8"),
                file_name=f"ranking_{nombre_proceso}.html",
                mime="text/html",
                help="Abre el archivo en tu navegador y usa Ctrl+P para imprimirlo o guardarlo como PDF.",
            )

            st.divider()
            st.markdown("## Detalle por candidato")
            for r in resultados:
                cv_json   = r["cv_json"]
                evaluacion = r["evaluacion"]
                nombre    = cv_json.get("nombre_candidato", "Desconocido")
                score     = evaluacion.get("score_total", 0)
                rec       = evaluacion.get("recomendacion", "descartar")

                with st.expander(f"{DECISION_COLOR.get(rec, '')} {nombre} — {score}/100"):
                    exp_col1, exp_col2 = st.columns([1, 1])

                    with exp_col1:
                        # Métricas de dimensión
                        dimensiones = [
                            ("🛠️ Skills",      evaluacion.get("score_skills_tecnicos")),
                            ("💼 Experiencia", evaluacion.get("score_experiencia")),
                            ("🎓 Educación",   evaluacion.get("score_educacion")),
                            ("🌐 Idiomas",     evaluacion.get("score_idiomas")),
                            ("🎯 Fit",         evaluacion.get("score_fit_general")),
                        ]
                        d_cols = st.columns(len(dimensiones))
                        for col, (label, val) in zip(d_cols, dimensiones):
                            col.metric(label, f"{val:.0f}/100" if val is not None else "—")

                        fortalezas = evaluacion.get("fortalezas", [])
                        brechas    = evaluacion.get("brechas", [])
                        st.success("**Puntos fuertes para este puesto:**\n\n" + "\n".join(f"• {f}" for f in fortalezas))
                        st.warning("**Aspectos a mejorar o que no cumplen el perfil:**\n\n" + "\n".join(f"• {b}" for b in brechas))

                        justificacion = evaluacion.get("justificacion_general", "")
                        st.info(f"**Por qué la IA tomó esta decisión:**\n\n{justificacion or 'No disponible.'}")
                        if justificacion:
                            st.caption("Haz clic en el ícono de copia para pegar directamente en un correo:")
                            st.code(justificacion, language=None)

                    with exp_col2:
                        st.plotly_chart(
                            radar_candidato(evaluacion, nombre),
                            use_container_width=True,
                        )

                    with st.expander("Ver perfil completo extraído del CV"):
                        st.json(cv_json)

# ==========================================
# TAB 3 — HISTORIAL
# ==========================================
with tab_historial:
    st.subheader("Historial de evaluaciones")

    if "historial_data" not in st.session_state:
        st.session_state.historial_data = []

    if st.button("Cargar / Actualizar historial", type="secondary"):
        with st.spinner("Consultando base de datos..."):
            try:
                resp = (
                    supabase.schema("gold")
                    .table("evaluaciones")
                    .select(
                        "id, score_total, recomendacion, created_at, cv_estructurado_id,"
                        "score_skills_tecnicos, score_experiencia, score_educacion,"
                        "score_idiomas, score_fit_general,"
                        "justificacion_general, tokens_input, tokens_output, costo_usd"
                    )
                    .order("created_at", desc=True)
                    .limit(200)
                    .execute()
                )
                evaluaciones = resp.data

                if not evaluaciones:
                    st.info("No hay evaluaciones registradas aún.")
                else:
                    ids = [e["cv_estructurado_id"] for e in evaluaciones]
                    resp_silver = (
                        supabase.schema("silver")
                        .table("cv_estructurados")
                        .select("id, nombre_candidato, ultimo_cargo, ultima_empresa, raw_cv_id")
                        .in_("id", ids)
                        .execute()
                    )
                    silver_map = {r["id"]: r for r in resp_silver.data}

                    raw_ids = [silver_map[i]["raw_cv_id"] for i in ids if i in silver_map]
                    resp_bronze = (
                        supabase.schema("bronze")
                        .table("raw_cvs")
                        .select("id, proceso_nombre")
                        .in_("id", raw_ids)
                        .execute()
                    )
                    bronze_map = {r["id"]: r["proceso_nombre"] for r in resp_bronze.data}

                    filas = []
                    for e in evaluaciones:
                        silver = silver_map.get(e["cv_estructurado_id"], {})
                        raw_id = silver.get("raw_cv_id")
                        filas.append({
                            "proceso_nombre":        bronze_map.get(raw_id, "—"),
                            "nombre_candidato":      silver.get("nombre_candidato", "—"),
                            "ultimo_cargo":          silver.get("ultimo_cargo", "—"),
                            "ultima_empresa":        silver.get("ultima_empresa", "—"),
                            "score_total":           e.get("score_total"),
                            "recomendacion":         e.get("recomendacion", "descartar"),
                            "score_skills_tecnicos": e.get("score_skills_tecnicos"),
                            "score_experiencia":     e.get("score_experiencia"),
                            "score_educacion":       e.get("score_educacion"),
                            "score_idiomas":         e.get("score_idiomas"),
                            "score_fit_general":     e.get("score_fit_general"),
                            "justificacion_general": e.get("justificacion_general", ""),
                            "tokens_input":          e.get("tokens_input") or 0,
                            "tokens_output":         e.get("tokens_output") or 0,
                            "costo_usd":             e.get("costo_usd") or 0.0,
                            "fecha":                 e["created_at"][:10] if e.get("created_at") else "—",
                        })

                    st.session_state.historial_data = filas

            except Exception as ex:
                st.error(f"Error al consultar Supabase: {ex}")

    if st.session_state.historial_data:
        filas = st.session_state.historial_data

        # --- Filtro por proceso ---
        procesos = sorted({f["proceso_nombre"] for f in filas if f["proceso_nombre"] != "—"})
        opciones = ["Todos los procesos"] + procesos
        filtro   = st.selectbox("Filtrar por proceso", opciones)
        filas_filtradas = filas if filtro == "Todos los procesos" else [
            f for f in filas if f["proceso_nombre"] == filtro
        ]

        st.caption(f"{len(filas_filtradas)} candidatos · {filtro}")

        # --- Dashboard: 3 gráficos ---
        st.markdown("### Resumen del proceso")
        g_col1, g_col2 = st.columns(2)

        with g_col1:
            st.markdown("**¿Cómo se distribuyeron las decisiones?**")
            st.plotly_chart(donut_decisiones(filas_filtradas), use_container_width=True)

        with g_col2:
            st.markdown("**¿Cómo se distribuyeron las puntuaciones?**")
            st.plotly_chart(histograma_scores(filas_filtradas), use_container_width=True)

        st.markdown("**Perfil promedio de los candidatos evaluados**")
        radar_col, _ = st.columns([1, 1])
        with radar_col:
            st.plotly_chart(radar_promedio(filas_filtradas), use_container_width=True)

        # --- Sección de costos de IA ---
        st.divider()
        st.markdown("### Costos del proceso de evaluación con IA")
        st.caption(
            "Cada CV requiere que la IA lea y analice el documento. "
            "Aquí puedes ver cuánto costó ese análisis, agrupado por proceso o por mes."
        )

        total_cvs   = len(filas_filtradas)
        total_costo = sum(f["costo_usd"] for f in filas_filtradas)

        kpi1, kpi2 = st.columns(2)
        kpi1.metric("CVs evaluados en esta selección", total_cvs)
        kpi2.metric("Costo total estimado (USD)", f"$ {total_costo:.4f}")

        agrupacion = st.radio(
            "Agrupar costos por:",
            ["Proceso de selección", "Mes"],
            horizontal=True,
            label_visibility="collapsed",
        )

        grupos: dict = {}
        for f in filas_filtradas:
            if agrupacion == "Proceso de selección":
                clave = f["proceso_nombre"]
            else:
                clave = mes_legible(f["fecha"]) if f["fecha"] != "—" else "Sin fecha"
            grupos.setdefault(clave, {"cvs": 0, "costo": 0.0})
            grupos[clave]["cvs"]   += 1
            grupos[clave]["costo"] += f["costo_usd"]

        tabla_costos = [
            {
                "Grupo":              g,
                "CVs evaluados":      v["cvs"],
                "Costo estimado (USD)": f"$ {v['costo']:.4f}",
            }
            for g, v in sorted(grupos.items())
        ]
        st.dataframe(tabla_costos, use_container_width=True, hide_index=True)

        # Gráfico de barras de costo por grupo
        if grupos:
            etiquetas = list(grupos.keys())
            costos    = [grupos[g]["costo"] for g in etiquetas]
            fig_costos = go.Figure(go.Bar(
                x=costos,
                y=etiquetas,
                orientation="h",
                marker_color="#6366f1",
                text=[f"$ {c:.4f}" for c in costos],
                textposition="outside",
            ))
            fig_costos.update_layout(
                xaxis=dict(title="Costo estimado (USD)"),
                yaxis=dict(autorange="reversed"),
                margin=dict(t=20, b=20, l=10, r=80),
                height=max(200, len(etiquetas) * 50),
            )
            st.plotly_chart(fig_costos, use_container_width=True)

        # --- Tabla de candidatos ---
        st.divider()
        st.markdown("### Lista de candidatos evaluados")
        tabla = [
            {
                "Proceso":     f["proceso_nombre"],
                "Candidato":   f["nombre_candidato"],
                "Último cargo": f["ultimo_cargo"],
                "Puntuación":  f["score_total"],
                "Decisión":    DECISION_ES.get(f["recomendacion"], f["recomendacion"].capitalize())
                               if f["recomendacion"] else "—",
                "Fecha":       f["fecha"],
            }
            for f in filas_filtradas
        ]
        st.dataframe(tabla, use_container_width=True, hide_index=True)
        st.caption(f"Mostrando {len(tabla)} registros.")

        # --- Exportación historial ---
        st.divider()
        st.markdown("### Exportar resultados")
        nombre_exp = filtro.replace(" ", "_") if filtro != "Todos los procesos" else "historial"
        h_col1, h_col2 = st.columns(2)

        buf_csv_h = io.StringIO()
        writer_h = csv.DictWriter(buf_csv_h, fieldnames=tabla[0].keys())
        writer_h.writeheader()
        writer_h.writerows(tabla)
        h_col1.download_button(
            label="⬇️ Descargar CSV",
            data=buf_csv_h.getvalue().encode("utf-8"),
            file_name=f"{nombre_exp}.csv",
            mime="text/csv",
            help="Archivo de texto separado por comas, compatible con Excel y Google Sheets.",
        )

        export_historial = [
            {
                "proceso_nombre":        f["proceso_nombre"],
                "nombre_candidato":      f["nombre_candidato"],
                "ultimo_cargo":          f["ultimo_cargo"],
                "ultima_empresa":        f["ultima_empresa"],
                "score_total":           f["score_total"],
                "score_skills_tecnicos": f["score_skills_tecnicos"],
                "score_experiencia":     f["score_experiencia"],
                "score_educacion":       f["score_educacion"],
                "score_idiomas":         f["score_idiomas"],
                "score_fit_general":     f["score_fit_general"],
                "recomendacion":         f["recomendacion"],
                "justificacion_general": f["justificacion_general"],
                "fecha":                 f["fecha"],
            }
            for f in filas_filtradas
        ]
        h_col2.download_button(
            label="📊 Descargar Excel completo",
            data=generar_excel(export_historial),
            file_name=f"{nombre_exp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Incluye justificaciones y colores por decisión. Ideal para presentar al equipo.",
        )
